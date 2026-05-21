"""Fetch vehicle registration statistics from Kraftfahrtbundesamt (KBA).

Source: FZ 3.1 — Bestand an Kraftfahrzeugen nach Gemeinden (xlsx).
Published annually; we fetch current year first, then fall back to previous year.

Attribution: Quelle: Kraftfahrtbundesamt (KBA)
License:     Datenlizenz Deutschland – Namensnennung – Version 2.0
"""
import logging
import os
from datetime import date
from io import BytesIO
from pathlib import Path

import httpx
import openpyxl

from fetchers.base import BaseFetcher
from models.schemas import VehiclesRaw

logger = logging.getLogger(__name__)

KBA_XLSX_BASE = (
    "https://www.kba.de/SharedDocs/Downloads/DE/Statistik/Fahrzeuge/FZ3/fz3_{year}.csv"
    "?__blob=publicationFile"
)
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "raw" / "kba"


def _normalise_ags(raw: str) -> str:
    """Extract and zero-pad 5-digit AGS from strings like '08115 BOEBLINGEN'."""
    raw = raw.strip()
    code = raw.split()[0] if raw else ""
    if not code.isdigit():
        return ""
    return code.zfill(5)


def _try_download(client: httpx.Client, year: int) -> bytes | None:
    url = KBA_XLSX_BASE.format(year=year)
    logger.info("KBA: trying %s", url)
    try:
        resp = client.get(url, timeout=60, follow_redirects=True)
        if resp.status_code == 404:
            logger.info("KBA: %d not found (404)", year)
            return None
        resp.raise_for_status()
        return resp.content
    except httpx.HTTPStatusError as exc:
        logger.warning("KBA: HTTP %s for year %d", exc.response.status_code, year)
        return None
    except Exception as exc:
        logger.warning("KBA: download failed for year %d: %s", year, exc)
        return None


def _parse_xlsx(raw_bytes: bytes) -> list[VehiclesRaw]:
    """Parse FZ 3.1 xlsx.

    Layout (confirmed from 2025 file):
      Sheet: 'FZ 3.1'
      Row 7 = primary headers, Row 8 = sub-headers (multi-level)
      Data starts row 9; col 1=Land, col 2=Zulassungsbezirk (5-digit AGS + name),
      col 3=Gemeinde or 'ZUSAMMEN' (Kreis aggregate), col 5=PKW insgesamt.

    We collect only 'ZUSAMMEN' rows (one per Zulassungsbezirk = Kreis aggregate)
    and track the current Zulassungsbezirk from the last non-None value in col 2.
    """
    wb = openpyxl.load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)

    # Find the data sheet
    ws = None
    for name in wb.sheetnames:
        if name.startswith("FZ 3"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    today = date.today()
    rows: list[VehiclesRaw] = []
    current_zb: str | None = None   # current Zulassungsbezirk cell value
    current_land: str | None = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 9:  # skip headers
            continue

        zb_cell = row[2]  # Zulassungsbezirk column (carries across Gemeinde rows)
        if zb_cell is not None:
            current_zb = str(zb_cell).strip()
        land_cell = row[1]
        if land_cell is not None:
            current_land = str(land_cell).strip()

        gemeinde_cell = row[3]
        if gemeinde_cell is None:
            continue
        gemeinde = str(gemeinde_cell).strip()
        if gemeinde != "ZUSAMMEN":
            continue

        # ZUSAMMEN row = Kreis-level aggregate
        if not current_zb:
            continue
        ags = _normalise_ags(current_zb)
        if not ags or len(ags) != 5:
            continue

        # Derive a name from the Zulassungsbezirk string (drop the code prefix)
        parts = current_zb.split(None, 1)
        district_name = parts[1].title() if len(parts) > 1 else current_zb

        try:
            cars_total = int(row[5]) if row[5] is not None else None
        except (ValueError, TypeError):
            cars_total = None

        rows.append(VehiclesRaw(
            ags=ags,
            district_name=district_name,
            cars_total=cars_total,
            cars_per_1000=None,
            electric=None,
            electric_share=None,
            hybrid=None,
            hybrid_share=None,
            data_date=today,
        ))

    wb.close()
    logger.info("KBA: parsed %d Kreis rows from xlsx", len(rows))
    return rows


class KbaFetcher(BaseFetcher):
    def fetch(self) -> list[VehiclesRaw]:
        os.makedirs(CACHE_DIR, exist_ok=True)
        today = date.today()
        current_year = today.year
        prev_year = current_year - 1

        raw_bytes: bytes | None = None
        with httpx.Client(timeout=60) as client:
            for year in (current_year, prev_year):
                raw_bytes = _try_download(client, year)
                if raw_bytes is not None:
                    logger.info("KBA: successfully downloaded FZ3 for year %d", year)
                    break

        if raw_bytes is None:
            logger.error("KBA: could not download FZ3 for years %d or %d", current_year, prev_year)
            return []

        cache_file = CACHE_DIR / f"{today.isoformat()}.xlsx"
        try:
            cache_file.write_bytes(raw_bytes)
            logger.info("KBA: cached xlsx to %s", cache_file)
        except OSError as exc:
            logger.warning("KBA: could not write cache: %s", exc)

        return _parse_xlsx(raw_bytes)

    def health_check(self) -> bool:
        year = date.today().year
        try:
            with httpx.Client(timeout=15) as client:
                resp = client.head(
                    KBA_XLSX_BASE.format(year=year),
                    follow_redirects=True,
                    timeout=15,
                )
                return resp.status_code in (200, 301, 302)
        except Exception:
            return False


def run() -> None:
    """Entry point called by scheduler and run_once.py."""
    from db.writer import write_vehicles_batch

    logger.info("=== KBA vehicle registrations fetch ===")
    fetcher = KbaFetcher()
    rows = fetcher.fetch()
    logger.info("KBA: writing %d rows to DB", len(rows))
    write_vehicles_batch(rows)
